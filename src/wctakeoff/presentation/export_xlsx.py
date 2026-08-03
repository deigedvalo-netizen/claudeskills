"""Spreadsheet export of the orderable takeoff (openpyxl 3.1.5).

Requirement ids: FR-7 (ADR-006, ADR-007, ADR-010).

Raw and with-waste columns are preserved side by side, each WC stays in
its own sale unit with no blended total, and the set-read reference
block and policy summary ship inside the workbook.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from wctakeoff.domain.units import UNKNOWN
from wctakeoff.presentation.presenter import TakeoffView, format_qty

_HEADER_FONT = Font(bold=True)


def write_xlsx(view: TakeoffView, dest: Path) -> Path:
    """Write the takeoff workbook: lines, set-read block, policy summary."""
    workbook = Workbook()

    lines = workbook.active
    lines.title = "Takeoff"
    lines.append(
        [
            "WC Tag",
            "Raw Qty (repeat loss incl.)",
            "With-Waste Qty",
            "Order Qty (rounded UP)",
            "Unit of Sale",
            "Contributing Walls",
        ]
    )
    for cell in lines[1]:
        cell.font = _HEADER_FONT
    for line_view in view.lines:
        line = line_view.line
        lines.append(
            [
                line.wc_tag,
                format_qty(line.raw_qty),
                format_qty(line.with_waste_qty),
                format_qty(line.order_qty),
                line.unit.value,  # per-WC unit, never converted (ADR-007)
                len(line.contributing),
            ]
        )

    setread = workbook.create_sheet("Set Read (verbatim)")
    setread.append(
        [
            "WC Tag",
            "Manufacturer",
            "Pattern",
            "Roll Width (in)",
            "Repeat (in)",
            "Match Type",
            "Unit of Sale",
            "Yield per Unit",
        ]
    )
    for cell in setread[1]:
        cell.font = _HEADER_FONT

    def _verbatim(value: object) -> str:
        if value is UNKNOWN:
            return "UNKNOWN"
        return getattr(value, "value", str(value))

    for defn in view.set_read_block:
        setread.append(
            [
                defn.wc_tag,
                _verbatim(defn.manufacturer),
                _verbatim(defn.pattern_name),
                _verbatim(defn.roll_width_in),
                _verbatim(defn.repeat_in),
                _verbatim(defn.match_type),
                defn.unit_of_sale.value,
                _verbatim(defn.yield_per_unit),
            ]
        )

    policy_sheet = workbook.create_sheet("Policy")
    policy = view.policy_summary
    policy_rows: list[tuple[str, str]] = [
        ("Waste %", format_qty(policy.waste_pct)),
        ("Deduct openings", str(policy.deduct_openings)),
        ("Opening threshold (sf)", format_qty(policy.opening_threshold_sf)),
        ("Trim allowance (in)", format_qty(policy.trim_allowance_in)),
    ]
    for wc_tag, waste in sorted(policy.per_wc_waste.items()):
        policy_rows.append((f"Waste % override [{wc_tag}]", format_qty(waste)))
    for label, value in policy_rows:
        policy_sheet.append([label, value])

    workbook.save(dest)
    return dest
